"""Combine each annotator's Task 1 + Task 2 CSV hand-off files into a single
color-coded Excel workbook with ONE sheet (one file per annotator) so
annotators don't have to juggle two files or two tabs. Rationale columns are
omitted -- only the validity/category judgments themselves are scored.

Usage:
    python3 15_build_annotator_workbooks.py

Reads:
    results/human_validation/task1_react_validity_annotator{1,2}.csv
    results/human_validation/task2_quality_category_annotator{1,2}.csv

Writes:
    results/human_validation/annotator1_validation.xlsx
    results/human_validation/annotator2_validation.xlsx
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = "results/human_validation"
SHEET_NAME = "ReACT Validation"

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONTEXT_FILL_A = PatternFill("solid", fgColor="DDEBF7")
CONTEXT_FILL_B = PatternFill("solid", fgColor="EAF2FA")
ANSWER_FILL_A = PatternFill("solid", fgColor="FFF2CC")
ANSWER_FILL_B = PatternFill("solid", fgColor="FFE699")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)

# (header, width, is_answer_col, dropdown_choices_or_None)
COLS = [
    ("item_id", 8, False, None),
    ("article_title", 28, False, None),
    ("source_pdf_path", 38, False, None),
    ("doi_or_url", 22, False, None),
    ("actionable", 40, False, None),
    ("impact", 40, False, None),
    ("evidence", 40, False, None),
    ("actionable_valid", 12, True, ["YES", "NO"]),
    ("evidence_valid", 12, True, ["YES", "NO"]),
    ("impact_valid", 12, True, ["YES", "NO", "N/A"]),
    ("sound_valid", 12, True, ["YES", "NO"]),
    ("precise_valid", 12, True, ["YES", "NO"]),
    ("cat_onboarding", 12, True, ["YES", "NO"]),
    ("cat_code_standards", 14, True, ["YES", "NO"]),
    ("cat_testing_qa", 12, True, ["YES", "NO"]),
    ("cat_community", 12, True, ["YES", "NO"]),
    ("cat_documentation", 14, True, ["YES", "NO"]),
    ("cat_governance", 12, True, ["YES", "NO"]),
    ("cat_security", 12, True, ["YES", "NO"]),
    ("cat_cicd", 10, True, ["YES", "NO"]),
]


def load_csv(path):
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def merge_rows(t1_rows, t2_rows):
    t2_by_id = {r["item_id"]: r for r in t2_rows}
    merged = []
    for r1 in t1_rows:
        r2 = t2_by_id.get(r1["item_id"], {})
        merged.append({**r1, **r2})
    return merged


def build_sheet(ws, rows):
    ws.title = SHEET_NAME
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False

    for c, (header, width, _, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30

    for r, row in enumerate(rows, start=2):
        ws.row_dimensions[r].height = 60
        band = r % 2 == 0
        for c, (header, _, is_answer, _) in enumerate(COLS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(header, ""))
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if is_answer:
                cell.fill = ANSWER_FILL_A if band else ANSWER_FILL_B
            else:
                cell.fill = CONTEXT_FILL_A if band else CONTEXT_FILL_B

    for c, (header, _, is_answer, choices) in enumerate(COLS, start=1):
        if choices:
            dv = DataValidation(
                type="list",
                formula1='"{}"'.format(",".join(choices)),
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            col_letter = get_column_letter(c)
            dv.add(f"{col_letter}2:{col_letter}{len(rows) + 1}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"


def build_legend(ws):
    ws.title = "Legend"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    rows = [
        ("Color", "Meaning"),
        ("Blue shading", "Context columns from the pipeline — read only, do not edit."),
        ("Yellow shading", "Your answer columns — fill these in (dropdowns where applicable)."),
        ("", ""),
        ("actionable_valid / evidence_valid / impact_valid",
         "Grounding check — is this actually supported by the source PDF? Needs source_pdf_path."),
        ("sound_valid / precise_valid / cat_*",
         "Quality and category check — judged from the actionable/impact/evidence text alone, no PDF needed."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        cb.alignment = WRAP
        if r == 1:
            for cell in (ca, cb):
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        elif a == "Blue shading":
            ca.fill = CONTEXT_FILL_A
        elif a == "Yellow shading":
            ca.fill = ANSWER_FILL_A


def build_workbook(annotator_num):
    t1_path = os.path.join(DIR, f"task1_react_validity_annotator{annotator_num}.csv")
    t2_path = os.path.join(DIR, f"task2_quality_category_annotator{annotator_num}.csv")
    t1_rows = load_csv(t1_path)
    t2_rows = load_csv(t2_path)
    rows = merge_rows(t1_rows, t2_rows)

    wb = Workbook()
    build_legend(wb.active)
    build_sheet(wb.create_sheet(), rows)

    out_path = os.path.join(DIR, f"annotator{annotator_num}_validation.xlsx")
    wb.save(out_path)
    print(f"wrote {out_path} ({len(rows)} rows)")


if __name__ == "__main__":
    for n in (1, 2):
        build_workbook(n)
