"""Reconcile the two independent annotators' labels (read directly from their
completed Excel workbooks), compute inter-annotator agreement (Cohen's
kappa) for both Task 1 (ReACT validity) and Task 2 (SOUND/PRECISE/category),
identify disagreements for the round-2 reconciliation step, merge the final
ground truth once round-2 reconciled labels are available, cross-check
Task 2 against the five-model-consensus pipeline's own verdicts to compute
Precision/Recall/F1, and write the results directly into the paper's LaTeX
tables.

Usage:
    python3 13_reconcile_human_validation.py round1
        -> reads results/human_validation/annotator{1,2}_validation.xlsx
           (Round-1, independent answers)
        -> prints + writes kappa report for Task 1 and Task 2
        -> writes results/human_validation/disagreements_for_round2.xlsx
           (single workbook, one sheet per task, both annotators' answers
           side by side, blank "final_*" columns for the reconciliation
           discussion)

    python3 13_reconcile_human_validation.py round2
        -> reads results/human_validation/disagreements_for_round2.xlsx
           (with the "final_*" columns filled in after annotators discussed
           every disagreement item)
        -> merges Round-1 agreements + Round-2 reconciled labels into the
           final ground truth for both tasks
        -> cross-checks Task 2 ground truth against
           results/human_validation/coordinator_key_200.csv (model consensus
           labels) to compute Precision/Recall/F1 per dimension
        -> writes results/human_validation/ground_truth_task{1,2}.csv
        -> regenerates -TOSEM-ReACT-LLM-v02/tables/human_validation_task1.tex
           and human_validation_task2.tex with the real numbers, and removes
           the "Status: pending" placeholder note from the manuscript
"""

import csv
import os
import re
import sys
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = "results/human_validation"
COORDINATOR_KEY_CSV = os.path.join(DIR, "coordinator_key_200.csv")
PAPER_DIR = "-TOSEM-ReACT-LLM-v02"
TABLE1_TEX = os.path.join(PAPER_DIR, "tables", "human_validation_task1.tex")
TABLE2_TEX = os.path.join(PAPER_DIR, "tables", "human_validation_task2.tex")
MAIN_TEX = os.path.join(PAPER_DIR, "01-main_no_template.tex")

VALIDATION_SHEET = "ReACT Validation"
TASK1_FIELDS = ["actionable_valid", "evidence_valid", "impact_valid"]
CAT_SLUGS = [
    "cat_onboarding", "cat_code_standards", "cat_testing_qa", "cat_community",
    "cat_documentation", "cat_governance", "cat_security", "cat_cicd",
]
CAT_LABELS = {
    "cat_onboarding": "New Contributor Onboarding and Involvement",
    "cat_code_standards": "Code Standards and Maintainability",
    "cat_testing_qa": "Automated Testing and Quality Assurance",
    "cat_community": "Community Collaboration and Engagement",
    "cat_documentation": "Documentation Practices",
    "cat_governance": "Project Management and Governance",
    "cat_security": "Security Best Practices and Legal Compliance",
    "cat_cicd": "CI/CD and DevOps Automation",
}
TASK2_FIELDS = ["sound_valid", "precise_valid"] + CAT_SLUGS

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONTEXT_FILL = PatternFill("solid", fgColor="DDEBF7")
ANSWER_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)


# ---------------------------------------------------------------- helpers --

def load_sheet(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        item_id = record.get("item_id")
        if item_id:
            rows[item_id] = {k: ("" if v is None else str(v).strip()) for k, v in record.items()}
    return rows


def cohens_kappa(labels_a, labels_b):
    n = len(labels_a)
    if n == 0:
        return None
    categories = sorted(set(labels_a) | set(labels_b))
    po = sum(1 for a, b in zip(labels_a, labels_b) if a == b) / n
    count_a = Counter(labels_a)
    count_b = Counter(labels_b)
    pe = sum((count_a[c] / n) * (count_b[c] / n) for c in categories)
    if pe == 1.0:
        return 1.0 if po == 1.0 else 0.0
    return (po - pe) / (1 - pe)


def interpret_kappa(k):
    if k is None:
        return "n/a"
    if k < 0:
        return "poor (worse than chance)"
    if k < 0.20:
        return "slight"
    if k < 0.40:
        return "fair"
    if k < 0.60:
        return "moderate"
    if k < 0.80:
        return "substantial"
    return "almost perfect"


def composite_valid(row):
    """A ReACT is 'valid' overall iff actionable and evidence are YES, and
    impact is YES or N/A (an action-only ReACT with no stated impact is not
    penalized for lacking one)."""
    if row["actionable_valid"] != "YES":
        return "NO"
    if row["evidence_valid"] != "YES":
        return "NO"
    if row["impact_valid"] not in ("YES", "N/A"):
        return "NO"
    return "YES"


def field_kappa(ann_a, ann_b, common_ids, field):
    pairs = [
        (ann_a[i][field], ann_b[i][field])
        for i in common_ids
        if ann_a[i][field] not in ("", "N/A") and ann_b[i][field] not in ("", "N/A")
    ]
    la = [p[0] for p in pairs]
    lb = [p[1] for p in pairs]
    k = cohens_kappa(la, lb)
    agree = sum(1 for a, b in zip(la, lb) if a == b)
    return {"n": len(pairs), "agree": agree, "kappa": k}


# ------------------------------------------------------------------ round1 --

def round1():
    a1 = load_sheet(os.path.join(DIR, "annotator1_validation.xlsx"), VALIDATION_SHEET)
    a2 = load_sheet(os.path.join(DIR, "annotator2_validation.xlsx"), VALIDATION_SHEET)
    # both tasks now live in the same single sheet/dict; kept as separate
    # names below only to minimize the diff against the task1/task2 split logic
    a1_t1, a2_t1, a1_t2, a2_t2 = a1, a2, a1, a2

    common_t1 = [i for i in a1_t1 if i in a2_t1]
    common_t2 = [i for i in a1_t2 if i in a2_t2]

    print(f"Task 1: {len(common_t1)}/{len(a1_t1)} items answered by both annotators")
    print(f"Task 2: {len(common_t2)}/{len(a1_t2)} items answered by both annotators\n")

    report_rows = []

    print("-- Task 1: ReACT validity --")
    for field in TASK1_FIELDS:
        r = field_kappa(a1_t1, a2_t1, common_t1, field)
        if r["n"] == 0:
            print(f"{field}: no answered items yet")
            continue
        pct = 100 * r["agree"] / r["n"]
        print(f"{field}: n={r['n']} raw_agreement={r['agree']}/{r['n']} ({pct:.1f}%) "
              f"kappa={r['kappa']:.3f} ({interpret_kappa(r['kappa'])})")
        report_rows.append(["task1", field, r["n"], r["agree"], round(pct, 1),
                             round(r["kappa"], 4), interpret_kappa(r["kappa"])])

    comp_a = {i: composite_valid(a1_t1[i]) for i in common_t1 if a1_t1[i]["actionable_valid"]}
    comp_b = {i: composite_valid(a2_t1[i]) for i in common_t1 if a2_t1[i]["actionable_valid"]}
    comp_ids = [i for i in comp_a if i in comp_b]
    if comp_ids:
        k_comp = cohens_kappa([comp_a[i] for i in comp_ids], [comp_b[i] for i in comp_ids])
        agree_comp = sum(1 for i in comp_ids if comp_a[i] == comp_b[i])
        pct = 100 * agree_comp / len(comp_ids)
        print(f"composite_valid: n={len(comp_ids)} raw_agreement={agree_comp}/{len(comp_ids)} "
              f"({pct:.1f}%) kappa={k_comp:.3f} ({interpret_kappa(k_comp)})")
        report_rows.append(["task1", "composite_valid", len(comp_ids), agree_comp, round(pct, 1),
                             round(k_comp, 4), interpret_kappa(k_comp)])

    print("\n-- Task 2: Quality and category --")
    for field in TASK2_FIELDS:
        r = field_kappa(a1_t2, a2_t2, common_t2, field)
        if r["n"] == 0:
            print(f"{field}: no answered items yet")
            continue
        pct = 100 * r["agree"] / r["n"]
        print(f"{field}: n={r['n']} raw_agreement={r['agree']}/{r['n']} ({pct:.1f}%) "
              f"kappa={r['kappa']:.3f} ({interpret_kappa(r['kappa'])})")
        report_rows.append(["task2", field, r["n"], r["agree"], round(pct, 1),
                             round(r["kappa"], 4), interpret_kappa(r["kappa"])])

    report_path = os.path.join(DIR, "round1_kappa_report.csv")
    with open(report_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["task", "field", "n_items", "raw_agreement", "pct_agreement",
                     "cohens_kappa", "interpretation"])
        w.writerows(report_rows)
    print(f"\nWrote {report_path}")

    write_disagreements_workbook(a1_t1, a2_t1, common_t1, a1_t2, a2_t2, common_t2)


def write_disagreements_workbook(a1_t1, a2_t1, common_t1, a1_t2, a2_t2, common_t2):
    disagree_t1 = [i for i in common_t1
                    if any(a1_t1[i][f] != a2_t1[i][f] for f in TASK1_FIELDS) and a1_t1[i]["actionable_valid"]]
    disagree_t2 = [i for i in common_t2
                    if any(a1_t2[i][f] != a2_t2[i][f] for f in TASK2_FIELDS) and a1_t2[i]["sound_valid"]]

    wb = Workbook()
    wb.remove(wb.active)

    def build(ws, title, ids, ann_a, ann_b, fields):
        ws.title = title
        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False
        headers = ["item_id"]
        for f in fields:
            headers += [f"A_{f}", f"B_{f}", f"final_{f}"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP_CENTER
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.column_dimensions["A"].width = 8
        for r, item_id in enumerate(ids, start=2):
            ws.row_dimensions[r].height = 22
            ws.cell(row=r, column=1, value=item_id).border = THIN_BORDER
            c = 2
            for f in fields:
                va = ann_a[item_id][f]
                vb = ann_b[item_id][f]
                ca = ws.cell(row=r, column=c, value=va); ca.fill = CONTEXT_FILL; ca.border = THIN_BORDER
                cb = ws.cell(row=r, column=c + 1, value=vb); cb.fill = CONTEXT_FILL; cb.border = THIN_BORDER
                cf = ws.cell(row=r, column=c + 2, value=""); cf.fill = ANSWER_FILL; cf.border = THIN_BORDER
                choices = ["YES", "NO", "N/A"] if "impact" in f else ["YES", "NO"]
                dv = DataValidation(type="list", formula1='"{}"'.format(",".join(choices)), allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(cf.coordinate)
                c += 3
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ids) + 1}"

    build(wb.create_sheet(), "Task 1 disagreements", disagree_t1, a1_t1, a2_t1, TASK1_FIELDS)
    build(wb.create_sheet(), "Task 2 disagreements", disagree_t2, a1_t2, a2_t2, TASK2_FIELDS)

    out_path = os.path.join(DIR, "disagreements_for_round2.xlsx")
    wb.save(out_path)
    print(f"\n{len(disagree_t1)}/{len(common_t1)} Task 1 items and {len(disagree_t2)}/{len(common_t2)} "
          f"Task 2 items disagree -> wrote {out_path} for the round-2 reconciliation discussion")


# ------------------------------------------------------------------ round2 --

def load_final_sheet(path, sheet_name, fields):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        item_id = record.get("item_id")
        if not item_id:
            continue
        final = {}
        flagged = False
        for f in fields:
            v = record.get(f"final_{f}")
            v = "" if v is None else str(v).strip()
            if not v:
                v = "NO" if "NO" in (record.get(f"A_{f}"), record.get(f"B_{f}")) else (record.get(f"A_{f}") or "NO")
                flagged = True
            final[f] = v
        out[item_id] = {"fields": final, "flag": flagged}
    return out


def round2():
    a1 = load_sheet(os.path.join(DIR, "annotator1_validation.xlsx"), VALIDATION_SHEET)
    a2 = load_sheet(os.path.join(DIR, "annotator2_validation.xlsx"), VALIDATION_SHEET)
    a1_t1, a2_t1, a1_t2, a2_t2 = a1, a2, a1, a2

    disagree_path = os.path.join(DIR, "disagreements_for_round2.xlsx")
    recon_t1 = load_final_sheet(disagree_path, "Task 1 disagreements", TASK1_FIELDS) if os.path.exists(disagree_path) else {}
    recon_t2 = load_final_sheet(disagree_path, "Task 2 disagreements", TASK2_FIELDS) if os.path.exists(disagree_path) else {}

    all_ids = [i for i in a1_t1 if a1_t1[i]["actionable_valid"] and i in a2_t1]

    gt_t1 = {}
    unresolved_t1 = 0
    for i in all_ids:
        if i in recon_t1:
            row = recon_t1[i]["fields"]
            if recon_t1[i]["flag"]:
                unresolved_t1 += 1
        else:
            row = {f: a1_t1[i][f] for f in TASK1_FIELDS}
        row["composite_valid"] = composite_valid(row)
        gt_t1[i] = row

    all_ids_t2 = [i for i in a1_t2 if a1_t2[i]["sound_valid"] and i in a2_t2]
    gt_t2 = {}
    unresolved_t2 = 0
    for i in all_ids_t2:
        if i in recon_t2:
            row = recon_t2[i]["fields"]
            if recon_t2[i]["flag"]:
                unresolved_t2 += 1
        else:
            row = {f: a1_t2[i][f] for f in TASK2_FIELDS}
        gt_t2[i] = row

    write_ground_truth_csvs(gt_t1, gt_t2)

    task1_stats = compute_task1_stats(a1_t1, a2_t1, gt_t1)
    task2_stats = compute_task2_stats(a1_t2, a2_t2, gt_t2)

    print(f"Task 1 ground truth: {len(gt_t1)} items ({unresolved_t1} defaulted to conservative label)")
    print(f"Task 2 ground truth: {len(gt_t2)} items ({unresolved_t2} defaulted to conservative label)")

    write_latex_tables(task1_stats, task2_stats)
    unmark_pending_status()


def write_ground_truth_csvs(gt_t1, gt_t2):
    path1 = os.path.join(DIR, "ground_truth_task1.csv")
    with open(path1, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["item_id"] + TASK1_FIELDS + ["composite_valid"])
        w.writeheader()
        for i, row in gt_t1.items():
            w.writerow({"item_id": i, **row})
    print(f"Wrote {path1}")

    path2 = os.path.join(DIR, "ground_truth_task2.csv")
    with open(path2, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["item_id"] + TASK2_FIELDS)
        w.writeheader()
        for i, row in gt_t2.items():
            w.writerow({"item_id": i, **row})
    print(f"Wrote {path2}")


def compute_task1_stats(a1, a2, gt):
    ids = list(gt.keys())
    stats = {}
    for field in TASK1_FIELDS:
        r = field_kappa(a1, a2, ids, field)
        stats[field] = r
    comp_a = {i: composite_valid(a1[i]) for i in ids}
    comp_b = {i: composite_valid(a2[i]) for i in ids}
    k_comp = cohens_kappa([comp_a[i] for i in ids], [comp_b[i] for i in ids])
    agree_comp = sum(1 for i in ids if comp_a[i] == comp_b[i])
    stats["composite_valid"] = {"n": len(ids), "agree": agree_comp, "kappa": k_comp}
    stats["pct_valid"] = {
        field: 100 * sum(1 for i in ids if gt[i][field] == "YES") / len(ids)
        for field in TASK1_FIELDS
    }
    stats["pct_composite_valid"] = 100 * sum(1 for i in ids if gt[i]["composite_valid"] == "YES") / len(ids)
    stats["n"] = len(ids)
    return stats


def compute_task2_stats(a1, a2, gt):
    ids = list(gt.keys())
    with open(COORDINATOR_KEY_CSV, newline="", encoding="utf-8") as fh:
        model = {r["item_id"]: r for r in csv.DictReader(fh)}

    stats = {"n": len(ids), "dims": {}}
    for field in TASK2_FIELDS:
        r = field_kappa(a1, a2, ids, field)
        model_field = {"sound_valid": "model_sound_label", "precise_valid": "model_precise_label"}.get(field, field)
        tp = fp = fn = tn = 0
        for i in ids:
            if i not in model:
                continue
            pred = model[i][model_field] == "YES"
            human = gt[i][field] == "YES"
            if pred and human:
                tp += 1
            elif pred and not human:
                fp += 1
            elif not pred and human:
                fn += 1
            else:
                tn += 1
        precision = tp / (tp + fp) if (tp + fp) else float("nan")
        recall = tp / (tp + fn) if (tp + fn) else float("nan")
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) and precision == precision and recall == recall else float("nan")
        stats["dims"][field] = {**r, "precision": precision, "recall": recall, "f1": f1,
                                  "tp": tp, "fp": fp, "fn": fn, "tn": tn}
    return stats


# --------------------------------------------------------------- LaTeX out --

def fmt(x, pct=False, digits=3):
    if x is None or x != x:  # None or NaN
        return "\\textcolor{red}{[PENDING]}"
    if pct:
        return f"{x:.1f}\\%"
    return f"{x:.{digits}f}"


def write_latex_tables(t1, t2):
    rows = []
    for field, label in [("actionable_valid", "Actionable"), ("evidence_valid", "Evidence"), ("impact_valid", "Impact")]:
        r = t1.get(field, {})
        n = r.get("n", 0)
        agree_pct = 100 * r["agree"] / r["n"] if r.get("n") else None
        rows.append((label, n, fmt(agree_pct, pct=True), fmt(r.get("kappa")),
                      fmt(t1.get("pct_valid", {}).get(field), pct=True)))
    comp = t1.get("composite_valid", {})
    comp_agree_pct = 100 * comp["agree"] / comp["n"] if comp.get("n") else None
    overall_row = ("Overall (all three components valid)", t1.get("n", 0),
                    fmt(comp_agree_pct, pct=True), fmt(comp.get("kappa")),
                    fmt(t1.get("pct_composite_valid"), pct=True))

    table1 = """\\begin{table}[!t]
\\centering
\\caption{Human Validation Pilot -- Task 1: ReACT Validity (Grounding in the Source Article).}
\\label{table:HumanValidationTask1}
\\def\\arraystretch{1.2}
\\begin{tabular}{lcccc}
\\hline
\\textbf{Component} & \\textbf{N} & \\textbf{Round-1 Raw} & \\textbf{Round-1} & \\textbf{Reconciled} \\\\
 & & \\textbf{Agreement} & \\textbf{Cohen's $\\kappa$} & \\textbf{\\% Valid} \\\\
\\hline
""" + "\n".join(f"{label} & {n} & {a} & {k} & {p} \\\\" for label, n, a, k, p in rows) + """
\\hline
""" + f"{overall_row[0]} & {overall_row[1]} & {overall_row[2]} & {overall_row[3]} & {overall_row[4]} \\\\" + """
\\hline
\\end{tabular}
\\end{table}
"""
    with open(TABLE1_TEX, "w", encoding="utf-8") as fh:
        fh.write(table1)
    print(f"Wrote {TABLE1_TEX}")

    dim_rows = []
    for field, label in [("sound_valid", "SOUND"), ("precise_valid", "PRECISE")]:
        d = t2["dims"].get(field, {})
        dim_rows.append((label, t2["n"], fmt(d.get("kappa")), fmt(d.get("precision")),
                          fmt(d.get("recall")), fmt(d.get("f1"))))
    cat_rows = []
    for slug in CAT_SLUGS:
        d = t2["dims"].get(slug, {})
        cat_rows.append((CAT_LABELS[slug], t2["n"], fmt(d.get("kappa")), fmt(d.get("precision")),
                          fmt(d.get("recall")), fmt(d.get("f1"))))

    table2 = """\\begin{table*}[!t]
\\centering
\\caption{Human Validation Pilot -- Task 2: Quality (SOUND/PRECISE) and Category Agreement.}
\\label{table:HumanValidationTask2}
\\def\\arraystretch{1.2}
\\resizebox{\\textwidth}{!}{
\\begin{tabular}{lccccc}
\\hline
\\textbf{Dimension} & \\textbf{N} & \\textbf{Round-1} & \\textbf{Human--Model} & \\textbf{Human--Model} & \\textbf{Human--Model} \\\\
 & & \\textbf{Cohen's $\\kappa$} & \\textbf{Precision} & \\textbf{Recall} & \\textbf{F1} \\\\
\\hline
""" + "\n".join(f"{label} & {n} & {k} & {p} & {r} & {f1} \\\\" for label, n, k, p, r, f1 in dim_rows) + """
\\hline
""" + "\n".join(f"{label} & {n} & {k} & {p} & {r} & {f1} \\\\" for label, n, k, p, r, f1 in cat_rows) + """
\\hline
\\end{tabular}
}
\\end{table*}
"""
    with open(TABLE2_TEX, "w", encoding="utf-8") as fh:
        fh.write(table2)
    print(f"Wrote {TABLE2_TEX}")


def unmark_pending_status():
    if not os.path.exists(MAIN_TEX):
        return
    with open(MAIN_TEX, encoding="utf-8") as fh:
        text = fh.read()
    pattern = re.compile(
        r"\\textcolor\{red\}\{\\textbf\{Status: pending\.\}.*?\}\n\n"
    )
    new_text, n = pattern.subn("", text, count=1)
    if n:
        with open(MAIN_TEX, "w", encoding="utf-8") as fh:
            fh.write(new_text)
        print(f"Removed the pending-status placeholder note from {MAIN_TEX}")


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else ""
    if mode == "round1":
        round1()
    elif mode == "round2":
        round2()
    else:
        print(__doc__)
        sys.exit(1)
